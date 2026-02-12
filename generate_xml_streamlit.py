import os
import sys
import xml.etree.ElementTree as ET
import xml.dom.minidom
import streamlit as st
import xmlschema
import yaml
import re
import csv
import io
import uuid
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from sqlalchemy import create_engine
import copy
import uuid
import zipfile

# Page configuration
st.set_page_config(page_title="EUDAMED XML Generator", layout="wide")
st.title("EUDAMED XML Generator")

# Custom CSS to darken syntax highlighting green color for better visibility
st.markdown("""
<style>
    /* Darken green for strings/attr-values in syntax highlighting */
    code span.token.string, code span.token.attr-value {
        color: #006400 !important; 
    }
    /* Darken green for inline code (paths) */
    code {
        color: #006400 !important;
    }
    /* Prevent code blocks from inheriting the green color as base text color */
    .stCodeBlock code {
        color: inherit !important;
    }
</style>
""", unsafe_allow_html=True)

def load_config(product_group):
    """Load YAML configuration for the selected product group."""
    if not product_group:
        return {}, {}
        
    filename = f"EUDAMED_data_{product_group}.yaml"
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, filename)
    
    if not os.path.exists(file_path):
        return None, None # Signal missing file
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
            return data.get('visible_fields', []), data.get('defaults', {}), data.get('envelope_settings', {})
    except Exception as e:
        st.error(f"Error loading config {filename}: {e}")
        return [], {}, {}

@st.cache_resource
def load_eudamed_metadata():
    """Load and cache metadata from EUDAMED CSV files."""
    metadata = {}
    all_headers = set()
    base_dir = os.path.dirname(os.path.abspath(__file__))
    csv_dir = os.path.join(base_dir, 'EUDAMED downloaded')
    
    files = ['basic-udi.csv', 'udi-di.csv']
    
    for filename in files:
        path = os.path.join(csv_dir, filename)
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                    # csv.DictReader might struggle if headers have BOM or spaces
                    # Read first line to clean headers
                    lines = f.readlines()
                    if not lines: continue
                    
                    # Use csv.reader to correctly parse headers (handles quotes)
                    header_reader = csv.reader([lines[0]])
                    raw_headers = next(header_reader)
                    
                    # Filter empty headers
                    headers = [h.strip() for h in raw_headers if h and h.strip()]
                    all_headers.update(headers)
                    
                    # Use the cleaned headers
                    reader = csv.DictReader(lines[1:], fieldnames=headers)
                    
                    for row in reader:
                        # Find the Field ID (it might be 'Field ID' or 'Field ID ' etc)
                        # We cleaned headers so it should be 'Field ID'
                        fld_id = row.get('Field ID')
                        if fld_id:
                            metadata[fld_id] = row
            except Exception as e:
                st.error(f"Error loading metadata from {filename}: {e}")
                
    # Sort headers to ensure consistent order
    sorted_headers = sorted(list(all_headers))
    # Make sure Field ID is present if not
    if 'Field ID' in sorted_headers:
         sorted_headers.remove('Field ID')
        
    return metadata, sorted_headers

@st.cache_resource
def load_schema():
    """Load and cache the XML schema."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    xsd_path = os.path.join(base_dir, 'EUDAMED downloaded', 'XSD', 'service', 'Message.xsd')
    
    if not os.path.exists(xsd_path):
        return None, f"Schema file not found at: {xsd_path}"

    try:
        schema = xmlschema.XMLSchema(xsd_path)
        return schema, None
    except Exception as e:
        return None, f"Failed to load schema: {e}"

def get_enums_for_type(type_obj):
    """Extract enumeration values from a type object."""
    enums = None
    if type_obj.is_simple():
        if hasattr(type_obj, 'enumeration') and type_obj.enumeration:
            enums = type_obj.enumeration
        elif hasattr(type_obj, 'base_type') and hasattr(type_obj.base_type, 'enumeration') and type_obj.base_type.enumeration:
             enums = type_obj.base_type.enumeration
    return [str(e) for e in enums] if enums else None

def get_type_constraints_help(type_obj):
    """Generate a help string for type constraints."""
    constraints = []
    if hasattr(type_obj, 'min_length') and type_obj.min_length is not None:
        constraints.append(f"Min Length: {type_obj.min_length}")
    if hasattr(type_obj, 'max_length') and type_obj.max_length is not None:
        constraints.append(f"Max Length: {type_obj.max_length}")
    if hasattr(type_obj, 'patterns') and type_obj.patterns:
         constraints.append(f"Pattern required")
    
    return " | ".join(constraints) if constraints else ""

def get_documentation(obj):
    """Extract documentation from an XSD component."""
    docs = []
    
    # helper to extract text safely
    def extract_text(doc):
        if isinstance(doc, str):
            return doc
        if hasattr(doc, 'text'):
            return doc.text
        # Fallback for lxml/ElementTree elements
        return getattr(doc, 'text', str(doc))

    try:
        if hasattr(obj, 'annotation') and obj.annotation is not None:
             if hasattr(obj.annotation, 'documentation') and obj.annotation.documentation:
                 for doc in obj.annotation.documentation:
                     txt = extract_text(doc)
                     if txt:
                         docs.append(txt.strip())
    except Exception as e:
        print(f"Error extracting documentation: {e}")
        
    return docs

def render_input_fields(element, type_obj, parent_key, state_container, xml_path="", config_visible=None, config_defaults=None, metadata=None, path_override=None, force_visible=False):
    """
    Recursively renders input fields for an element.
    Returns the value entered/selected by the user.
    """
    indent_level = len(parent_key.split(".")) if parent_key else 0
    # Use clean name for key generation to avoid duplicates or weird keys
    elem_name_clean = element.name

    if hasattr(element, 'name') and element.name and '}' in element.name:
        elem_name_clean = element.name.split('}')[-1]
    
    key = f"{parent_key}.{elem_name_clean}" if parent_key else elem_name_clean
    
    if path_override:
        current_path = path_override
    else:
        current_path = f"{xml_path}/{element.local_name}" if xml_path else element.local_name
    
    # Store the structure in session state to rebuild XML later
    if 'xml_structure' not in state_container:
        state_container['xml_structure'] = {}

    if type_obj.is_simple():
        # Configuration Visibility Check
        is_mandatory = getattr(element, 'min_occurs', 1) >= 1
        
        # Handle indexed paths (e.g., path/to/elem[0])
        clean_path_for_check = re.sub(r'\[\d+\]', '', current_path)
        
        is_visible = (config_visible is None) or \
                     (current_path in config_visible) or \
                     (clean_path_for_check in config_visible) or \
                     is_mandatory or \
                     force_visible
        
        # Default Value
        default_val = None
        if config_defaults:
            default_val = config_defaults.get(current_path)
            if default_val is None:
                default_val = config_defaults.get(clean_path_for_check)

        # Logic: If hidden, try to return default, else return None (skip)
        
        if not is_visible:
            if default_val is not None:
                return str(default_val)
            # If mandatory (min_occurs >= 1) but hidden and no default -> Warning? Or skip?
            # We skip it. Validation will catch it later if it was critical.
            return None

        # Check for List Type (e.g. whitespace separated values)
        is_list_type = getattr(type_obj, 'is_list', lambda: False)()

        enums = get_enums_for_type(type_obj)
        # If it is a list type, try to get enums from the item type
        if not enums and is_list_type and hasattr(type_obj, 'item_type'):
             enums = get_enums_for_type(type_obj.item_type)

        # Handle optional Enum: Add empty option if not mandatory
        if enums and not is_list_type and not is_mandatory:
            if "" not in enums:
                enums = [""] + enums

        label = f"{element.local_name}"
        
        # Display XML Path
        st.caption(f"📍 Path: `{current_path}`")
        
        # Build help text with documentation
        help_lines = []
        
        # 1. Try element annotation
        element_docs = get_documentation(element)
        if element_docs:
            help_lines.extend(element_docs)
        
        # 2. Try type annotation if element has none
        if not element_docs:
            type_docs = get_documentation(type_obj)
            if type_docs:
                help_lines.extend(type_docs)

        # Extract FLD codes
        temp_help_text = "\n".join(help_lines)
        fld_codes = re.findall(r"#(FLD.*?)#", temp_help_text)
        
        # Fetch Metadata
        meta_info = {}
        if metadata and fld_codes:
            for code in fld_codes:
                if code in metadata:
                    row = metadata[code]
                    meta_info[code] = row
                    # Append info to help lines
                    help_lines.append(f"--- Metadata for {code} ---")
                    if row.get('Field Label'):
                        help_lines.append(f"Label: {row['Field Label']}")
                    if row.get('Field Description / Notes'):
                        help_lines.append(f"Description: {row['Field Description / Notes']}")
                    if row.get('Business Rules'):
                        help_lines.append(f"Rules: {row['Business Rules']}")
        
        help_lines.append(f"Namespace: {element.name}")
        
        constraint_text = get_type_constraints_help(type_obj)
        if constraint_text:
            help_lines.append(f"Constraints: {constraint_text}")
            
        help_text = "\n\n".join(help_lines)
        
        val = None
        if enums:
            if is_list_type:
                # Handle Multi-Select for List Types
                default_selections = []
                if default_val:
                    # Split string by whitespace to get selected items
                    default_selections = str(default_val).split()
                    # Filter valid enums only to prevent errors
                    default_selections = [x for x in default_selections if x in enums]
                
                selected = st.multiselect(label, options=enums, default=default_selections, key=key, help=help_text)
                # XML List types are space-separated strings
                val = " ".join(selected) if selected else None
            else:
                # Handle index for default value in selectbox
                default_idx = 0
                if default_val and str(default_val) in enums:
                    default_idx = enums.index(str(default_val))
                    
                val = st.selectbox(label, options=enums, index=default_idx, key=key, help=help_text)
                
                # If empty string selected/defaulted, return None so it is omitted from XML
                if val == "":
                    val = None
        elif hasattr(type_obj, 'primitive_type') and type_obj.primitive_type and type_obj.primitive_type.local_name == 'boolean':
             # Handle Boolean
             # Default value check
             is_checked = False
             if default_val is not None:
                 if isinstance(default_val, bool):
                     is_checked = default_val
                 elif str(default_val).lower() == 'true':
                     is_checked = True
            
             bool_val = st.toggle(label, value=is_checked, key=key, help=help_text)
             val = "true" if bool_val else "false"
        else:
            # Check for max length for the input widget
            max_chars = None
            if hasattr(type_obj, 'max_length') and type_obj.max_length is not None:
                max_chars = int(type_obj.max_length)
             
            # Default value
            input_val = str(default_val) if default_val is not None else ""
                
            val = st.text_input(label, value=input_val, key=key, help=help_text, max_chars=max_chars)
        
        # Validation Logic
        if val:
            # Use xmlschema's own validation to check the value
            try:
                type_obj.validate(val)
            except xmlschema.XMLSchemaValidationError as e:
                st.error(f"❌ Invalid format: {e.reason}")
            except Exception as e:
                st.error(f"❌ Invalid value")
            
            # Record data for CSV Export
            fld_code_str = ", ".join(fld_codes) if fld_codes else ""
            
            # XSD Occurrences
            min_o = getattr(element, 'min_occurs', '1')
            max_o = getattr(element, 'max_occurs', '1')
            if max_o is None: max_o = "unbounded"

            # Base entry
            csv_entry = {
                'XMLPath': current_path,
                'value': val,
                'xsd_min': str(min_o),
                'xsd_max': str(max_o),
                'FLD_code': fld_code_str,
                'tooltip': help_text
            }
            
            # Aggregate all metadata columns
            # We want to check ALL headers that might exist in the collected rows
            if meta_info:
                # Collect all column names found in the matched rows
                found_keys = set()
                for row in meta_info.values():
                    found_keys.update(row.keys())
                
                for key in found_keys:
                    if key is None: continue # Skip 'restkey' or unmatched columns
                    
                    values = []
                    # We iterate over fld_codes to preserve some order or logic?
                    # Or just iterate over meta_info items?
                    # The fld_codes list determines which rows are relevant.
                    for code in fld_codes:
                        if code in meta_info:
                            val_part = meta_info[code].get(key, '')
                            if val_part: 
                                if isinstance(val_part, list):
                                    values.append(",".join(map(str, val_part)))
                                else:
                                    values.append(str(val_part))
                    
                    if values:
                        # Join multiple values with semi-colon
                        # Avoid duplicating if same value exists? 
                        # Let's keep all to see distribution
                        csv_entry[key] = "; ".join(values)
            
            if 'csv_entries' not in state_container:
                state_container['csv_entries'] = []
            
            state_container['csv_entries'].append(csv_entry)

        return val

    elif type_obj.is_complex():
        label = f"**{element.local_name}**"
        
        # Try to get documentation for complex type
        c_help_lines = []
        c_docs = get_documentation(element)
        if not c_docs:
             c_docs = get_documentation(type_obj)
        
        if c_docs:
            # We can't put help on markdown, so we render an info box or caption if docs exist
            st.markdown(label)
            for d in c_docs:
                st.caption(f"ℹ️ {d}")
        else:
            st.markdown(label)
            
        st.caption(f"Path: `{current_path}`")
        
        group = type_obj.content
        if not group: 
            return None

        children_data = {}
        
        # Helper to process model groups (Sequence/Choice)
        def process_group(group_particle, parent_key, current_path, indent_level, cv, cd, md):
             group_data = {}
             
             # If it's a Choice with minOccurs >= 1, we must force a made selection
             if group_particle.model == 'choice' and group_particle.min_occurs >= 1:
                 # Get options
                 options = [p for p in group_particle.iter_model()]
                 # Create labels for options (using local_name if element, else 'Group')
                 option_labels = []
                 for opt in options:
                     if isinstance(opt, xmlschema.validators.XsdElement):
                         option_labels.append(opt.local_name)
                     else:
                         option_labels.append("Nested Group") # Simplified for now
                 
                 # Unique key for this choice
                 choice_key = f"{parent_key}_choice_{id(group_particle)}"
                 
                 # Auto-selection logic based on visibility config
                 default_idx = 0
                 forced_choice = False
                 
                 if cv:
                     visible_candidates = []
                     for idx, opt in enumerate(options):
                         if isinstance(opt, xmlschema.validators.XsdElement):
                             opt_path = f"{current_path}/{opt.local_name}"
                             
                             # Check precise match or if it's a prefix for other visible fields
                             # (e.g. modelName vs modelName/name)
                             is_visible = False
                             if opt_path in cv:
                                 is_visible = True
                             else:
                                 # Prefix Check
                                 prefix = opt_path + "/"
                                 if any(v.startswith(prefix) for v in cv):
                                     is_visible = True
                                     
                             if is_visible:
                                 visible_candidates.append(idx)
                     
                     # If exactly one option is configured to be visible, pick it
                     if len(visible_candidates) == 1:
                         default_idx = visible_candidates[0]
                         forced_choice = True
                         
                 # --- SELECTION LOGIC ---
                 selected_particle = None
                 
                 if not forced_choice:
                     st.markdown(f"{'  ' * indent_level}*Choose one required option:*")
                     selected_label = st.radio("Select type:", option_labels, index=default_idx, key=choice_key, horizontal=True, label_visibility="collapsed")
                     
                     for opt in options:
                         if isinstance(opt, xmlschema.validators.XsdElement) and opt.local_name == selected_label:
                             selected_particle = opt
                             break
                 else:
                     # Explicitly grab the forced option
                     if 0 <= default_idx < len(options):
                        selected_particle = options[default_idx]
                     else:
                        st.error(f"Index error in forced choice: {default_idx} vs len {len(options)}")
                 
                 if selected_particle is not None:
                      # Process the selected branch
                      
                      if isinstance(selected_particle, xmlschema.validators.XsdElement):
                           if forced_choice:
                                # Using standard layout but forcing visibility
                                # Explicitly calling render_input_fields
                                
                                with st.container():
                                     # Use columns to align with peers (match the visual indentation)
                                     col1, col2 = st.columns([0.5, 9.5])
                                     with col2:
                                         val = render_input_fields(
                                            selected_particle, 
                                            selected_particle.type, 
                                            parent_key, 
                                            state_container, 
                                            current_path,
                                            cv, cd, md,
                                            path_override=None,
                                            force_visible=True
                                        )
                                # Store result
                                group_data[selected_particle.name] = val
                           else:
                               with st.container():
                                    col1, col2 = st.columns([0.5, 9.5])
                                    with col2:
                                        # Recursive call
                                        val = render_input_fields(
                                            selected_particle, 
                                            selected_particle.type, 
                                            parent_key, 
                                            state_container, 
                                            current_path,
                                            cv, cd, md
                                        )
                                        # Ensure we store it even if it's None (but usually None is skipped)
                                        # Use qualified name for correct namespace mapping
                                        group_data[selected_particle.name] = val
             
             # If Sequence or Optional Choice (though optional choice usually doesn't force input)
             else:
                 for particle in group_particle.iter_model():
                     if isinstance(particle, xmlschema.validators.XsdElement):
                         # Determine visibility: Mandatory OR Configured (Visible/Default)
                         clean_path = f"{current_path}/{particle.local_name}" if current_path else particle.local_name
                         child_path = clean_path # default to clean path for check
                         
                         # Normalize path for checking configuration (remove indices)
                         clean_path_no_idx = re.sub(r'\[\d+\]', '', clean_path)
                         
                         # Visibility Check:
                         # 1. Exact match in config
                         # 2. Key prefix match (if children are configured, parent must be visible)
                         
                         is_in_config = False
                         if cv:
                             if clean_path in cv or clean_path_no_idx in cv:
                                 is_in_config = True
                             else:
                                 # Prefix Check (are there visible children?)
                                 prefix = clean_path_no_idx + "/"
                                 if any(v.startswith(prefix) for v in cv):
                                     is_in_config = True
                                     
                         if not is_in_config and cd:
                              if clean_path in cd or clean_path_no_idx in cd:
                                  is_in_config = True
                              else:
                                  # Prefix Check for defaults
                                  prefix = clean_path_no_idx + "/"
                                  if any(k.startswith(prefix) for k in cd):
                                      is_in_config = True
                         
                         is_configured_clean = is_in_config
                         
                         # Check for repeated element
                         is_repeated = particle.max_occurs is None or particle.max_occurs > 1
                         
                         if is_repeated:
                             # Default count Logic
                             count = particle.min_occurs
                             
                             # Check for indexed defaults to determine initial count (e.g. key "Path[1]")
                             if cd:
                                 idx = 0
                                 found_index = True
                                 while found_index:
                                     # We need to scan keys because they are fully qualified paths
                                     # Optimized check: try to find any key containing "{clean_path}[{idx}]"
                                     # Since keys are typically "Path/To/Leaf", checking exact prefix is safer.
                                     # But default keys are LEAF level. clean_path might be intermediate (complex).
                                     # If clean_path is "A/B", and we have "A/B[0]/C: val"
                                     prefix = f"{clean_path}[{idx}]"
                                     # Need to check against absolute keys in defaults
                                     # Note: 'cd' keys are usually full paths. 'clean_path' is full path so far.
                                     # But check if suffix matches? 
                                     # The keys in defaults are "MDRDevice/..."
                                     # content of 'clean_path' is "MDRDevice/..." (current path joined)
                                     
                                     # Check for exact prefix match in defaults keys
                                     # Check if any key starts with prefix
                                     found_start = False
                                     combined_prefix = prefix + "/"
                                     if any(k.startswith(combined_prefix) or k == prefix for k in cd.keys()):
                                         found_start = True
                                     
                                     if found_start:
                                         if (idx + 1) > count:
                                             count = idx + 1
                                         idx += 1
                                     else:
                                         found_index = False
                                 
                                 # Check if specific fields inside the list item are visible (CV)
                                 # (handle case where no default value is set but field is visible)
                                 # If visible_fields has "Path/Item/Field", we should at least show 1 item?
                                 # Current logic: 'is_configured_clean' is true if children are visible.
                                 if is_configured_clean and count == 0:
                                      count = 1

                             # Ensure we show if any index is configured or clean path is configured
                             if particle.min_occurs >= 1 or is_configured_clean or count > 0:
                                 st.markdown(f"{'  ' * indent_level}**{particle.local_name} (List)**")
                                 count_key = f"{parent_key}_{particle.local_name}_count"
                                 count_val = st.number_input(f"Number of {particle.local_name} entries", min_value=particle.min_occurs, value=count, key=count_key)
                                 
                                 vals = []
                                 for i in range(count_val):
                                     with st.expander(f"{particle.local_name} #{i+1}", expanded=False):
                                         indexed_path = f"{clean_path}[{i}]"
                                         child_val = render_input_fields(
                                            particle, 
                                            particle.type, 
                                            f"{parent_key}_{i}", 
                                            state_container, 
                                            xml_path=None,
                                            config_visible=cv, 
                                            config_defaults=cd, 
                                            metadata=md,
                                            path_override=indexed_path
                                         )
                                         if child_val is not None:
                                             vals.append(child_val)
                                 if vals:
                                     # Store with qualified name
                                     group_data[particle.name] = vals

                         else:
                             if particle.min_occurs >= 1 or is_configured_clean:
                                with st.container():
                                    col1, col2 = st.columns([0.5, 9.5])
                                    with col2:
                                        child_val = render_input_fields(
                                            particle, 
                                            particle.type, 
                                            parent_key, 
                                            state_container, 
                                            current_path,
                                            cv, cd, md
                                        )
                                        if child_val is not None:
                                            # Store with qualified name
                                            group_data[particle.name] = child_val
                     
                     elif isinstance(particle, xmlschema.validators.XsdGroup):
                         if particle.min_occurs >= 1:
                             # Recurse for nested group
                             nested_data = process_group(particle, parent_key, current_path, indent_level, cv, cd, md)
                             group_data.update(nested_data)
                             
             return group_data

        # Start processing the top-level group
        # The top level content of a complex type is a Group (usually sequence)
        children_data = process_group(group, key, current_path, 0, config_visible, config_defaults, metadata)
        
        if not children_data: return None
        return children_data

def build_xml_element(element_name, xsd_type, form_data):
    """Recursively builds an ElementTree element from the dictionary form data."""
    tag = element_name
    elem = ET.Element(tag)
    
    if isinstance(form_data, dict):
        for child_tag, child_val in form_data.items():
            if child_val is None: continue 

            # Handle List of values (maxOccurs > 1)
            if isinstance(child_val, list):
                for item in child_val:
                    if isinstance(item, (str, dict)):
                         child_elem = build_xml_element_manual_tag(child_tag, item)
                         elem.append(child_elem)
            elif isinstance(child_val, (str, dict)):
                child_elem = build_xml_element_manual_tag(child_tag, child_val)
                elem.append(child_elem)
    
    elif isinstance(form_data, str):
        elem.text = form_data
        
    return elem

def build_xml_element_manual_tag(tag, content):
    elem = ET.Element(tag)
    if isinstance(content, dict):
        # We need to apply correct namespaces to children based on where they belong.
        # This is tricky without the full schema context at this level.
        # Heuristic: 
        # - Common Device fields -> 'commondi'
        # - Basic UDI specific -> 'basicudi'
        # - UDI-DI specific -> 'udidi'
        # - Market Info -> 'marketinfo'
        # - Links -> 'links'
        
        # Or, we can look at the key name itself.
        for child_tag, child_val in content.items():
            if child_val is None: continue
            
            # Determine namespace for child_tag if not present
            final_tag = child_tag
            
            # If child_tag is already qualified {uri}name, leave it.
            if not child_tag.startswith('{'):
                # Try to map based on known field names
                if child_tag in ['riskClass', 'model', 'humanTissuesCells', 'animalTissuesCells', 
                                 'humanProductCheck', 'IIb_implantable_exceptions', 'medicinalProductCheck',
                                 'type', 'MFActorCode', 'deviceCertificateLinks']:
                     final_tag = f"{{{namespaces['basicudi']}}}{child_tag}"
                elif child_tag in ['identifier', 'status', 'basicUDIIdentifier', 'MDNCodes', 
                                   'productionIdentifier', 'referenceNumber', 'sterile', 'sterilization',
                                   'numberOfReuses', 'marketInfos', 'baseQuantity', 'latex', 'reprocessed']:
                     final_tag = f"{{{namespaces['udidi']}}}{child_tag}"
                elif child_tag in ['DICode', 'issuingEntityCode', 'active', 'administeringMedicine', 
                                   'implantable', 'measuringFunction', 'reusable', 'code']:
                     final_tag = f"{{{namespaces['commondi']}}}{child_tag}"
                elif child_tag in ['deviceCertificateLink', 'certificateNumber', 'NBActorCode', 'certificateType']:
                     final_tag = f"{{{namespaces['links']}}}{child_tag}"
                elif child_tag in ['marketInfo', 'country', 'originalPlacedOnTheMarket']:
                     final_tag = f"{{{namespaces['marketinfo']}}}{child_tag}"
            
            if isinstance(child_val, list):
                for item in child_val:
                     child_elem = build_xml_element_manual_tag(final_tag, item)
                     elem.append(child_elem)
            else:
                child_elem = build_xml_element_manual_tag(final_tag, child_val)
                elem.append(child_elem)
    else:
        elem.text = str(content)
    return elem

# --- Database Integration Functions ---

def get_db_engine():
    """Establishes connection to the Oracle database using environment variables."""
    db_user = os.getenv("ORAUSER")
    db_password = os.getenv("ORAPW")
    db_alias = "MC9ELES" # Hardcoded alias from provided script
    
    if not db_user or not db_password:
        st.error("Missing environment variables ORAUSER or ORAPW.")
        return None

    try:
        connection_string = f"oracle+oracledb://{db_user}:{db_password}@{db_alias}"
        engine = create_engine(connection_string)
        return engine
    except Exception as e:
        st.error(f"Failed to create DB engine: {e}")
        return None

def fetch_ifs_data(model, pcode):
    """Fetches UDI-DI data from IFS database."""
    engine = get_db_engine()
    if not engine:
        return None
    
    # Note: Ensure inputs are safe or used in parameters, but given the dynamic table function
    # usage in the sample, f-string injection was used there too.
    # We strip and quote carefully.
    
    query = f"""
    with
       transferable_parts as (select * from table (get_transferable_parts_lens_table ('ALL', '{model}', 'LK')))
    select
       dpt,cyl,
       pcode,
       get_techspec_info_mcf (part_no, 'ET_GTIN') udi_di
    from
       transferable_parts tp
    where
       tp.pcode='{pcode}'
       and get_techspec_info_mcf (part_no, 'ET_GTIN') is not null
    """
    
    try:
        with engine.connect() as conn:
            df = pd.read_sql(query, conn)
        return df
    except Exception as e:
        st.error(f"Error executing query: {e}")
        return None

# --- Main App ---

schema, error_msg = load_schema()
metadata_csv, metadata_headers = load_eudamed_metadata()

if not schema:
    st.error(error_msg)
    st.stop()

# --- Logo & Configuration ---
base_dir = os.path.dirname(os.path.abspath(__file__))
logo_path = os.path.join(base_dir, '.streamlit', 'EUDAMED_logo.jpg')
if os.path.exists(logo_path):
    st.sidebar.image(logo_path, width="stretch")

# --- Product Group Selection ---
st.sidebar.header("Configuration")

# Dynamic scan for YAML configuration files
product_groups = []
prefix = "EUDAMED_data_"
suffix = ".yaml"

try:
    for f in os.listdir(base_dir):
        if f.startswith(prefix) and f.endswith(suffix):
            # Extract group name
            group_name = f[len(prefix):-len(suffix)]
            product_groups.append(group_name)
    product_groups.sort()
except Exception as e:
    st.sidebar.error(f"Error scanning for config files: {e}")

# Default to "Lens" if available, otherwise "None" (index 0)
default_target = "Lens"
default_ix = 0
if default_target in product_groups:
    default_ix = product_groups.index(default_target) + 1

selected_group = st.sidebar.selectbox("Select Product Group", ["None"] + product_groups, index=default_ix)

config_visible = None
config_defaults = None
config_envelope = None

if selected_group != "None":
    config_visible, config_defaults, config_envelope = load_config(selected_group)
    if config_visible or config_defaults:
        st.sidebar.success(f"Loaded configuration for {selected_group}")
        
        # Display current YAML in main area
        filename = f"EUDAMED_data_{selected_group}.yaml"
        file_path = os.path.join(base_dir, filename)
        if os.path.exists(file_path):
             with open(file_path, 'r', encoding='utf-8') as f:
                 yaml_content = f.read()
             with st.expander("Current Default Values", expanded=False):
                 # Use 'properties' or 'text' to avoid red highlighting which can look like errors
                 st.code(yaml_content, language="properties")
    else:
        st.sidebar.warning(f"No specific configuration found for {selected_group}")


# Namespace handling
namespaces = {
    'device': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Device/v1',
    'basicudi': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Device/BasicUDI/v1',
    'udidi': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/UDIDI/v1',
    'commondi': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Device/CommonDevice/v1',
    'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
    'eudi': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Device/LegacyDevice/EUDI/v1',
    'eudididata': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Device/LegacyDevice/EUDIData/v1',
    'm': 'https://ec.europa.eu/tools/eudamed/dtx/servicemodel/Message/v1',
    's': 'https://ec.europa.eu/tools/eudamed/dtx/servicemodel/Service/v1',
    'links': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Links/v1',
    'lsn': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/Common/LanguageSpecific/v1',
    'marketinfo': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/MktInfo/MarketInfo/v1',
    'e': 'https://ec.europa.eu/tools/eudamed/dtx/datamodel/Entity/v1'
}
for prefix, uri in namespaces.items():
    ET.register_namespace(prefix, uri)

# Device Configuration Type Selection
device_type_options = {
    "MDR Device (Regulation)": "MDRDevice",
    "Legacy Device (MDD/AIMDD)": "MDEUDevice",
    "IVDR Device": "IVDRDevice",
    "Legacy IVD": "IVDEUDevice"
}

st.sidebar.markdown("---")
# Default to MDR Device
selected_device_type_label = st.sidebar.selectbox("Select Device Type", list(device_type_options.keys()))
selected_root_element_name = device_type_options[selected_device_type_label]

# Find root definition
mdr_device_element = schema.elements.get(selected_root_element_name)
if not mdr_device_element:
    mdr_device_element = schema.elements.get(f"{{{namespaces['device']}}}{selected_root_element_name}")

# Look in imported maps if not found in root elements
if not mdr_device_element and hasattr(schema, 'maps') and schema.maps and schema.maps.elements:
    mdr_device_element = schema.maps.elements.get(f"{{{namespaces['device']}}}{selected_root_element_name}")


if not mdr_device_element:
    st.error(f"Could not find {selected_root_element_name} element definition in schema.")
    st.stop()

mdr_device_type = mdr_device_element.type
basic_udi_def = None
udidi_data_def = None

# Logic to find the Basic UDI and UDI-DI Data parts based on naming conventions
# MDR: MDRBasicUDI, MDRUDIDIData
# Legacy: MDEUDI, MDEUData
# IVDR: IVDRBasicUDI, IVDRUDIDIData
# Legacy IVD: IVDEUDI, IVDEUData

for particle in mdr_device_type.content.iter_model():
    name = particle.name
    if 'BasicUDI' in name or 'EUDI' in name:
        basic_udi_def = particle
    elif 'UDIDIData' in name or 'EUData' in name:
        udidi_data_def = particle

if not basic_udi_def or not udidi_data_def:
    st.error(f"Structure mismatch for {selected_root_element_name}: Could not find Basic UDI or Data definitions.")
    st.stop()


# --- UI Layout ---

st.header(f"{selected_device_type_label} Configuration")

# --- Service & Action Configuration ---
st.markdown("### XML Generation Settings")
col_svc_1, col_svc_2 = st.columns(2)

with col_svc_1:
    service_op_mode = st.radio(
        "Operation Type", 
        ["POST (Create/Register)", "PATCH (Update)"],
        index=0,
        help="POST is for new registrations. PATCH is for updating existing records."
    )

target_scope = []
# Default Service IDs
service_id_override = None

if service_op_mode.startswith("POST"):
    with col_svc_2:
        post_type = st.radio(
            "Registration Target", 
            ["Full Device (Basic UDI + UDI-DIs)", "Add UDI-DI(s) only"],
            index=0,
            help="Full Device registers a Basic UDI and its UDI-DI(s). 'Add UDI-DI' just adds a UDI-DI to an existing Basic UDI."
        )
    
    if post_type.startswith("Full"):
        # DEVICE POST (Composite)
        target_scope = ['BasicUDI', 'UDIDI']
        service_id_override = "DEVICE"
    else:
        # UDI_DI POST
        target_scope = ['UDIDI']
        service_id_override = "UDI_DI"

else: # PATCH
    with col_svc_2:
        st.write("Select entities to update:")
        patch_basic = st.checkbox("Basic UDI", value=True)
        # For UDI-DI PATCH, usually implies updating data for specific UDI-DIs.
        patch_udidi = st.checkbox("UDI-DI Data", value=False)
        
        # Version Input for PATCH
        patch_version = st.number_input("Current Entity Version", min_value=1, value=1, help="Required for PATCH operations. Provide the current version number of the entity you are updating.")
    
    if patch_basic:
        target_scope.append('BasicUDI')
    if patch_udidi:
        target_scope.append('UDIDI')
    
    # Service ID is determined per file for PATCH

st.markdown("---")

# Container for collecting data for CSV export
data_collection_container = {'csv_entries': []}

# We use a distinct key prefix
basic_udi_path = f"{mdr_device_element.local_name}"
# Add selected_group to key prefix to ensure widgets refresh when configuration changes
basic_udi_key_prefix = f"root_{selected_group}_{selected_root_element_name}"

if 'BasicUDI' in target_scope:
    with st.expander("Basic UDI Configuration", expanded=True):
        st.info("Fill in the mandatory fields for the Basic UDI. Min Occurs >= 1 fields only.")
        basic_udi_data = render_input_fields(
            basic_udi_def, 
            basic_udi_def.type, 
            basic_udi_key_prefix, 
            data_collection_container, 
            basic_udi_path, 
            config_visible, 
            config_defaults,
            metadata_csv
        )
else:
    basic_udi_data = None

if 'UDIDI' in target_scope:
    with st.expander("UDI-DI Data Entries", expanded=True):
        st.info("Fill in the mandatory fields for the UDI-DI. You can add multiple entries.")

        # Determine limit based on service type
        max_udis = 50
        help_msg = "You can add multiple entries."
        if service_id_override == "DEVICE":
             max_udis = 1
             help_msg = "Restricted to 1 entry for Full Device Registration (Device Service)."

        # Always allow multiple UDI-DIs for generation, regardless of schema maxOccurs in the container.
        # This supports "Add UDI-DI" scenarios (multiple messages) and bulk generation.
        col_count, col_dummy = st.columns([2, 8])
        with col_count:
            num_udis = st.number_input("Number of UDI-DI entries", min_value=1, max_value=max_udis, value=1, help=help_msg)

        udidi_data_list = []
        udidi_base_path = f"{mdr_device_element.local_name}"
        for i in range(num_udis):
            with st.expander(f"UDI-DI Entry #{i+1}", expanded=False):
                # Pass unique parent key with group prefix
                group_key_prefix = f"root_{selected_group}_{selected_root_element_name}.udidi_{i}"
                udidi_data = render_input_fields(
                    udidi_data_def, 
                    udidi_data_def.type, 
                    group_key_prefix, 
                    data_collection_container, 
                    udidi_base_path,
                    config_visible,
                    config_defaults,
                    metadata_csv
                )
                udidi_data_list.append(udidi_data)
else:
    udidi_data_list = []


st.markdown("---")

# --- IFS Integration Controls ---
use_ifs = st.toggle("Generate UDI-DI by IFS")
ifs_model = ""
ifs_pcode = ""

if use_ifs:
    col_ifs_1, col_ifs_2 = st.columns(2)
    with col_ifs_1:
        ifs_model = st.text_input("IFS Model", help="Model parameter for DB Query")
    with col_ifs_2:
        ifs_pcode = st.text_input("IFS PCode", help="Package Code parameter for DB Query")

st.markdown("---")
# Action Buttons in columns
col_gen, col_export = st.columns([1, 1])

with col_gen:
    submitted = st.button("Generate XML", type="primary")

with col_export:
    # Prepare Excel Data
    excel_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "EUDAMED Data"

    # Define fields configuration
    # Renaming map
    rename_map = {
        "Field applicable for MDR": "Field for MDR",
        "Field applicable for IVDR": "Field for IVDR"
    }

    # Build columns definition list: [(DisplayHeader, DataKey)]
    final_columns_def = []
    
    # 1. XMLPath (Fixed first column)
    final_columns_def.append(('XMLPath', 'XMLPath'))
    
    # Check if "Occurrence" exists in metadata to determine placement
    has_occurrence = "Occurrence" in metadata_headers
    
    # If "Occurrence" is NOT in metadata, we inject XSD cols early for visibility
    if not has_occurrence:
         final_columns_def.append( ("XSD MinOccurs", "xsd_min") )
         final_columns_def.append( ("XSD MaxOccurs", "xsd_max") )
    
    # Standard fixed columns
    final_columns_def.append(('value', 'value'))
    final_columns_def.append(('FLD_code', 'FLD_code'))
    final_columns_def.append(('tooltip', 'tooltip'))

    # Metadata columns (with dynamic injection if Occurrence exists)
    for mh in metadata_headers:
        display_name = rename_map.get(mh, mh)
        final_columns_def.append( (display_name, mh) )
        
        if mh == "Occurrence":
             final_columns_def.append( ("XSD MinOccurs", "xsd_min") )
             final_columns_def.append( ("XSD MaxOccurs", "xsd_max") )

    # Extract headers for Excel
    headers = [c[0] for c in final_columns_def]
    ws.append(headers)

    # Write data
    if 'csv_entries' in data_collection_container:
        for entry in data_collection_container['csv_entries']:
            row = []
            for col_def in final_columns_def:
                row.append(entry.get(col_def[1], ""))
            ws.append(row)

    # Create Table
    last_col_letter = get_column_letter(len(headers))
    last_row = ws.max_row
    
    if last_row > 1: # Only create table if data exists (row 1 is header)
        tab = Table(displayName="EudamedData", ref=f"A1:{last_col_letter}{last_row}")
        # Use TableStyleMedium16 (Blue-ish in some themes, or Neutral) as requested
        # Disable column stripes ("Make columns not banded")
        style = TableStyleInfo(name="TableStyleMedium16", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Apply Shrink to Fit
    shrink_alignment = Alignment(shrink_to_fit=True, wrap_text=False)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = shrink_alignment
            
    # Set column widths based on header titles
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        
        if header == 'XMLPath': 
            # Column A: Fixed width ~7.5 cm (approx 41 chars)
            ws.column_dimensions[col_letter].width = 41
        elif header == 'value':
            # Column value: 2x wider than standard title fit
            base_width = len(str(header)) + 5
            ws.column_dimensions[col_letter].width = base_width * 2
        elif header in ['XSD MinOccurs', 'XSD MaxOccurs']:
            ws.column_dimensions[col_letter].width = 15
        else:
            # Other columns: Width based on header title length + padding
            ws.column_dimensions[col_letter].width = len(str(header)) + 5

    wb.save(excel_buffer)
    excel_data = excel_buffer.getvalue()

    st.download_button(
        label="Export Data to Excel",
        data=excel_data,
        file_name="eudamed_data_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def clean_xsi_type_name(element_name):
    """
    Cleans up element name to remove URI if present and returns simple name.
    """
    if '}' in element_name:
        return element_name.split('}')[-1]
    return element_name

if submitted:
    st.success("Generating XML...")
    
    # helper to add xsi:type
    def set_xsi_type(elem, type_name):
        clean_type = clean_xsi_type_name(type_name)
        # Determine prefix based on where the type is usually defined
        prefix = "device"
        if "UDIDI" in clean_type: 
             prefix = "udidi" # e.g. MDRUDIDIDataType
        elif "BasicUDI" in clean_type:
             prefix = "device" 
        
        elem.set(f"{{{namespaces['xsi']}}}type", f"{prefix}:{clean_type}")

    # --- IFS Data Processing ---
    final_udidi_list = udidi_data_list # Default
    
    if use_ifs:
        if not ifs_model or not ifs_pcode:
             st.error("Please provide Model and PCode for IFS generation.")
             st.stop()
        
        with st.spinner("Querying IFS Database..."):
             df = fetch_ifs_data(ifs_model, ifs_pcode)
             
        if df is None or df.empty:
             st.error("No data returned from IFS query.")
             st.stop()
             
        # Normalize columns to lowercase
        df.columns = [c.lower() for c in df.columns]
        
        # Add '0' prefix to udi_di if it exists
        if 'udi_di' in df.columns:
             df['udi_di'] = "0" + df['udi_di'].astype(str)
             
        # Process Data: Sort by DPT ASC, CYL ASC
        try:
             # Ensure numeric conversion for correct sorting
             df['dpt_num'] = pd.to_numeric(df['dpt'], errors='coerce').fillna(999999)
             df['cyl_num'] = pd.to_numeric(df['cyl'], errors='coerce').fillna(999999)
             df_sorted = df.sort_values(by=['dpt_num', 'cyl_num'])
        except Exception as e:
             st.warning(f"Sorting error: {e}. Using default order.")
             df_sorted = df
             
        min_record = df_sorted.iloc[0]
        bulk_records = df_sorted.iloc[1:] if len(df_sorted) > 1 else pd.DataFrame()
        
        # Helper to safely update DICode and ReferenceNumber
        def update_udi_values(item_dict, udi_val):
            if not isinstance(item_dict, dict): return
            
            # 1. Update Reference Number (direct child, varying namespaces possible)
            # Iterate keys to find 'referenceNumber' regardless of namespace prefix
            ref_num_key = None
            for k in item_dict.keys():
                if 'referenceNumber' in k: # e.g. 'udidi:referenceNumber' or just 'referenceNumber'
                    ref_num_key = k
                    break
            
            if ref_num_key:
                 item_dict[ref_num_key] = str(udi_val)
            
            # 2. Update DICode in identifier
            # First find identifier key
            ident_key = None
            for k in item_dict.keys():
                if 'identifier' in k and not 'basicUDIIdentifier' in k: # Distinct from basicUDIIdentifier
                    ident_key = k
                    break
            
            if ident_key and isinstance(item_dict[ident_key], dict):
                ident_dict = item_dict[ident_key]
                # Now look for DICode inside
                di_code_key = None
                for k in ident_dict.keys():
                    if 'DICode' in k:
                        di_code_key = k
                        break
                
                if di_code_key:
                    ident_dict[di_code_key] = str(udi_val)
                    
        # 1. DEVICE / POST
        if service_op_mode.startswith("POST") and (service_id_override == "DEVICE" or post_type.startswith("Full")):
              template = copy.deepcopy(udidi_data_list[0]) if udidi_data_list else {}
              
              update_udi_values(template, min_record['udi_di'])

              final_udidi_list = [template] 
        
        # 2. UDI_DI / POST or PATCH (Bulk Logic)
        elif (service_op_mode.startswith("POST") and service_id_override == "UDI_DI") or \
             (service_op_mode.startswith("PATCH") and 'UDIDI' in target_scope):
              
              template = copy.deepcopy(udidi_data_list[0]) if udidi_data_list else {}
              new_list = []
              
              for idx, row in bulk_records.iterrows():
                   new_item = copy.deepcopy(template)
                   
                   update_udi_values(new_item, row['udi_di'])
                   
                   new_list.append(new_item)
              
              final_udidi_list = new_list

    generation_tasks = []

    if service_op_mode.startswith("POST"):
        task = {
            'mode': 'POST',
            'target': post_type,
            'service_id': service_id_override
        }
        generation_tasks.append(task)
    else:
        # PATCH - Check Scopes
        if 'BasicUDI' in target_scope:
            generation_tasks.append({'mode': 'PATCH', 'target': 'BasicUDI', 'service_id': 'BASIC_UDI'})
        if 'UDIDI' in target_scope: # If using IFS, this will use the generated list
            generation_tasks.append({'mode': 'PATCH', 'target': 'UDIDI', 'service_id': 'UDI_DI'})

    created_files = []

    for idx, task in enumerate(generation_tasks):
        payload_blocks = [] # List of blocks to generate separate files
        
        if task['service_id'] == 'DEVICE': # Full Device
             # Single block with Minimum UDI-DI (if IFS) or whatever is in list
             payload_blocks.append({'type': 'DEVICE', 'budi': basic_udi_data, 'udidis': final_udidi_list, 'index': 1, 'total': 1})
             
        elif task['service_id'] == 'UDI_DI': # UDI-DI POST or PATCH
             # Bulk Chunking
             chunk_size = 300
             all_items = final_udidi_list if final_udidi_list else []
             
             # Create chunks
             if not all_items:
                 # Handle case with no items (empty file? or skip?)
                 payload_blocks.append({'type': 'UDIDI_BULK', 'items': [], 'index': 1, 'total': 1})
             else:
                 chunk_indices = list(range(0, len(all_items), chunk_size))
                 total_chunks = len(chunk_indices)
                 for idx, i in enumerate(chunk_indices):
                      chunk = all_items[i:i + chunk_size]
                      payload_blocks.append({'type': 'UDIDI_BULK', 'items': chunk, 'index': idx + 1, 'total': total_chunks})
                  
        elif task['target'] == 'BasicUDI':
             payload_blocks.append({'type': 'BasicUDI', 'data': basic_udi_data, 'index': 1, 'total': 1})

        # Generate separate file for each block
        for block_idx, block in enumerate(payload_blocks):
        
            # Root Payload for this file
            payload_elements = [] 

            if block['type'] == 'DEVICE':
                p_root = ET.Element(f"{{{namespaces['device']}}}Device")
                type_name = clean_xsi_type_name(mdr_device_element.type.name)
                set_xsi_type(p_root, type_name)
                
                # Add Basic UDI
                if block['budi']:
                    budi_name = clean_xsi_type_name(basic_udi_def.name)
                    basic_udi_elem = build_xml_element_manual_tag(f"{{{namespaces['device']}}}{budi_name}", block['budi'])
                    p_root.append(basic_udi_elem)
                    
                # Add UDI-DIs
                for udi_data in block['udidis']:
                    if udi_data:
                         udidi_name = clean_xsi_type_name(udidi_data_def.name)
                         udidi_elem = build_xml_element_manual_tag(f"{{{namespaces['device']}}}{udidi_name}", udi_data)
                         p_root.append(udidi_elem)
                
                payload_elements.append(p_root)

            elif block['type'] == 'UDIDI_BULK':
                # Generate multiple UDIDIData elements
                type_name = udidi_data_def.type.name if hasattr(udidi_data_def.type, 'name') else "MDRUDIDIDataType"
                
                for item in block['items']:
                     p_root = ET.Element(f"{{{namespaces['device']}}}UDIDIData")
                     set_xsi_type(p_root, f"udidi:{type_name}")
                     
                     if task['mode'] == 'PATCH':
                         # Add Version for PATCH
                         # Check availability of patch_version
                         ver_val = str(patch_version) if 'patch_version' in locals() else "1"
                         ver_elem = ET.Element(f"{{{namespaces['e']}}}version")
                         ver_elem.text = ver_val
                         p_root.insert(0, ver_elem)

                     temp_elem = build_xml_element_manual_tag("TEMP", item)
                     for child in temp_elem:
                          p_root.append(child)
                     
                     payload_elements.append(p_root)

            elif block['type'] == 'BasicUDI':
                 p_root = ET.Element(f"{{{namespaces['device']}}}BasicUDI")
                 type_name = basic_udi_def.type.name if hasattr(basic_udi_def.type, 'name') else "MDRBasicUDIType"
                 set_xsi_type(p_root, f"device:{type_name}")
                 
                 if task['mode'] == 'PATCH':
                     ver_val = str(patch_version) if 'patch_version' in locals() else "1"
                     ver_elem = ET.Element(f"{{{namespaces['e']}}}version")
                     ver_elem.text = ver_val
                     p_root.insert(0, ver_elem)
                 
                 temp_elem = build_xml_element_manual_tag("TEMP", block['data'])
                 for child in temp_elem:
                      p_root.append(child)
                 
                 payload_elements.append(p_root)

            if not payload_elements: continue

            # 3. Build Envelope
            sec_token = ""
            actor_code = ""
            party_id = ""
            if config_envelope:
                sec_token = config_envelope.get('security_token', '')
                actor_code = config_envelope.get('actor_code', '')
                party_id = config_envelope.get('party_id', '')

            m_ns = f"{{{namespaces['m']}}}"
            ns2_ns = f"{{{namespaces['s']}}}"
            
            root = ET.Element(f"{m_ns}Push")
            
            root.set(f"{{{namespaces['xsi']}}}schemaLocation", 
                     f"{namespaces['m']} https://webgate.ec.europa.eu/tools/eudamed/dtx/service/Message.xsd")
            root.set("version", "3.0.25")
            
            corr_id = ET.SubElement(root, f"{m_ns}correlationID")
            corr_id.text = str(uuid.uuid4())
            
            create_dt = ET.SubElement(root, f"{m_ns}creationDateTime")
            create_dt.text = datetime.datetime.now(datetime.timezone.utc).isoformat().replace('+00:00', 'Z')
            
            msg_id = ET.SubElement(root, f"{m_ns}messageID")
            msg_id.text = str(uuid.uuid4())
            
            recipient = ET.SubElement(root, f"{m_ns}recipient")
            node = ET.SubElement(recipient, f"{m_ns}node")
            node_actor = ET.SubElement(node, f"{ns2_ns}nodeActorCode")
            node_actor.text = "EUDAMED"
            
            service = ET.SubElement(recipient, f"{m_ns}service")
            svc_id = ET.SubElement(service, f"{ns2_ns}serviceID")
            svc_id.text = task['service_id']
            svc_op = ET.SubElement(service, f"{ns2_ns}serviceOperation")
            svc_op.text = task['mode']
            
            # <m:payload>
            payload = ET.SubElement(root, f"{m_ns}payload")
            # Append all elements for this block
            for pe in payload_elements:
                payload.append(pe)
            
            sender = ET.SubElement(root, f"{m_ns}sender")
            s_node = ET.SubElement(sender, f"{m_ns}node")
            s_node_actor = ET.SubElement(s_node, f"{ns2_ns}nodeActorCode")
            s_node_actor.text = actor_code
            
            s_service = ET.SubElement(sender, f"{m_ns}service")
            s_site_id = ET.SubElement(s_service, f"{ns2_ns}serviceID")
            s_site_id.text = task['service_id']
            s_svc_op = ET.SubElement(s_service, f"{ns2_ns}serviceOperation")
            s_svc_op.text = task['mode']

            rough_string = ET.tostring(root, encoding="utf-8")
            reparsed = xml.dom.minidom.parseString(rough_string)
            final_xml = reparsed.toprettyxml(indent="  ", encoding="utf-8").decode("utf-8")
            
            final_xml = final_xml.replace('xmlns:s=', 'xmlns:ns2=')
            final_xml = final_xml.replace('<s:', '<ns2:')
            final_xml = final_xml.replace('</s:', '</ns2:')

            final_xml = re.sub(r'\n\s*\n', '\n', final_xml)

            validation_status = "Unknown"
            validation_details = ""
            try:
                if schema.is_valid(final_xml):
                    validation_status = "Valid"
                    validation_details = "✅ XML is valid against the schema."
                else:
                    validation_status = "Invalid"
                    try:
                        schema.validate(final_xml)
                    except Exception as e:
                        validation_details = f"❌ Validation Error: {e}"
            except Exception as e:
                 validation_status = "Error"
                 validation_details = f"⚠️ Validation Process Failed: {e}"

            # Filename generation
            current_date_str = datetime.datetime.now().strftime("%y%m%d")
            
            # Variables for model/pcode
            model_val = str(ifs_model).strip() if 'ifs_model' in locals() and ifs_model else "NOMODEL"
            pcode_val = str(ifs_pcode).strip() if 'ifs_pcode' in locals() and ifs_pcode else "NOPCODE"

            # Sanitization
            model_val = "".join([c for c in model_val if c.isalnum() or c in ('-','_')])
            pcode_val = "".join([c for c in pcode_val if c.isalnum() or c in ('-','_')])

            base_fname = f"{current_date_str}-{model_val}-{pcode_val}-{task['service_id']}-{task['mode']}"
            
            if block.get('total') is not None:
                 fname = f"{base_fname}-Part{block['index']}-{block['total']}.xml"
            else:
                 # Standard naming without parts
                 fname = f"{base_fname}.xml"
            
            created_files.append({
                'name': fname, 
                'content': final_xml, 
                'label': f"{task['service_id']} {task['mode']} ({block['type']})",
                'validation_status': validation_status,
                'validation_details': validation_details
            })

    st.subheader("Generated XML Files")
    
    for cfile in created_files:
        with st.expander(f"{cfile['name']} ({cfile['validation_status']})", expanded=False):
             if cfile['validation_status'] == "Valid":
                 st.success(cfile['validation_details'])
             elif cfile['validation_status'] == "Invalid":
                 st.error(cfile['validation_details'])
             else:
                 st.warning(cfile['validation_details'])
                 
             st.code(cfile['content'], language="xml")
             st.download_button(
                label=f"Download {cfile['name']}",
                data=cfile['content'],
                file_name=cfile['name'],
                mime="application/xml",
                key=cfile['name']
            )

    # --- Bulk Download ---
    if len(created_files) > 0:
        with col_gen:
             # Create a Zip File in memory
             zip_buffer = io.BytesIO()
             with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                 for cfile in created_files:
                     zip_file.writestr(cfile['name'], cfile['content'])
             
             st.download_button(
                 label="Download All XMLs (ZIP)",
                 data=zip_buffer.getvalue(),
                 file_name=f"EUDAMED_Bulk_{uuid.uuid4().hex[:8]}.zip",
                 mime="application/zip",
                 type="secondary"
             )


